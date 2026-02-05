import os

import pandas as pd
from openpyxl.styles import Font

def calc_total(top1_5_avg_df, pr_avg_df, conf_df, top1_5_avg_df_subopt, pr_avg_df_subopt):
    """
    Calculate averages for all metrics except confusion matrix for the optimal data
    """
    top1_5_total_df = top1_5_avg_df.groupby("name", as_index=False).mean()
    pr_total_df = pr_avg_df.groupby("name", as_index=False).mean()
    top1_5_total_df_subopt = top1_5_avg_df_subopt.groupby("name", as_index=False).mean()
    pr_total_df_subopt = pr_avg_df_subopt.groupby("name", as_index=False).mean()

    conf_total_dfs = {}
    for alg_dfs in conf_df:
        conf_total_dfs[alg_dfs]  = pd.concat(conf_df[alg_dfs].values()).groupby(level=0).sum()

    return top1_5_total_df, pr_total_df, conf_total_dfs, top1_5_total_df_subopt, pr_total_df_subopt

def write_dfs_to_sheet(
    writer: pd.ExcelWriter,

    dfs: dict,                 # {title: DataFrame}
    sheet_name: str,
    startrow: int = 0,
    space: int = 2,
    bold_titles: bool = True,
) -> int:
    """
    Write multiple DataFrames to the same Excel sheet, each with a title above it.
    Works with engine="openpyxl".
    Returns the next available row (0-based).
    """
    wb = writer.book  # openpyxl Workbook
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.create_sheet(title=sheet_name)
    # Ensure it's visible (prevents IndexError on save in some edge cases)
    ws.sheet_state = "visible"
    # Register worksheet in pandas writer
    writer.sheets[sheet_name] = ws
    row = startrow
    # If empty dict, still write something so the sheet exists and is visible
    if not dfs:
        cell = ws.cell(row=row + 1, column=1, value="No tables to write (dfs is empty).")
        if bold_titles:
            cell.font = Font(bold=True)
        return row + 1
    for title, df in dfs.items():
        # Title (openpyxl uses 1-based row/col)
        cell = ws.cell(row=row + 1, column=1, value=str(title))
        if bold_titles:
            cell.font = Font(bold=True)
        # DataFrame below title
        df.to_excel(
            writer,
            sheet_name=sheet_name,
            startrow=row + 1,   # pandas is 0-based
            startcol=0,
            index=False,
        )
        # Advance pointer:
        # 1 row for title + (header + len(df)) rows for dataframe + space rows
        row += 1 + (len(df) + 1) + space
    return row


if __name__ == "__main__":
    """
    In results_info the details of the runs are indicated : folder id is the key, and the values are a list containing
     fold number, experiment description
    """
    networks_path = '/media/bella/8A1D-C0A6/Academy/Home_ultrasound/output/network'
    results_path = '/home/bella/Academy/results/Opt_cases/overall_summary.xlsx'
    sheets_names = ["top1-5", 'precision_recall', 'conf_matrix']
    top_column_names = ['wfss', 'wfss_top3', 'wfss_top5', 'accuracy', 'accuracy_top3', 'accuracy_top5', 'inference_time']
    top_column_names_subopt = ['accuracy', 'accuracy_top3', 'accuracy_top5', 'inference_time']
    pr_column_names = ['precision_opt_only', 'recall_opt_only', 'F1_opt_only',	'precision_opt_subopt',
                       'recall_opt_subopt', 'F1_opt_subopt']
    pr_column_names_subopt = ['precision_opt_subopt', 'recall_opt_subopt', 'F1_opt_subopt']
    conf_column_names = ['irrelevant', 'optimal',	'suboptimal']
    conf_column_names_subopt = ['irrelevant', 'suboptimal']
    # results_info = {170: [0, 'opt only\wo mixup densenet'],
    #                 171: [1, 'opt only\wo mixup densenet'],
    #                 172: [2, 'opt only\wo mixup densenet'],
    #                 173: [3, 'opt only\wo mixup densenet'],
    #                 174: [4, 'opt only\wo mixup densenet'],
    #                 206: [0, 'Three classes Densenet \wo mixup '],
    #                 207: [1, 'Three classes Densenet \wo mixup '],
    #                 208: [2, 'Three classes Densenet \wo mixup '],
    #                 209: [3, 'Three classes Densenet \wo mixup '],
    #                 210: [4, 'Three classes Densenet \wo mixup '],
    #                 196: [0, 'One-step Densenet \wo mixup'],
    #                 197: [1, 'One-step Densenet \wo mixup'],
    #                 198: [2, 'One-step Densenet \wo mixup'],
    #                 199: [3, 'One-step Densenet \wo mixup'],
    #                 200: [4, 'One-step Densenet \wo mixup'],
    #                 201: [0, 'two-step Densenet \wo mixup'],
    #                 202: [1, 'two-step Densenet \wo mixup'],
    #                 203: [2, 'two-step Densenet \wo mixup'],
    #                 204: [3, 'two-step Densenet \wo mixup'],
    #                 205: [4, 'two-step Densenet \wo mixup'],
    #                 }
    results_info = {
                    201: [0, 'two-step Densenet \wo mixup'],
                    202: [1, 'two-step Densenet \wo mixup'],
                    203: [2, 'two-step Densenet \wo mixup'],
                    204: [3, 'two-step Densenet \wo mixup'],
                    205: [4, 'two-step Densenet \wo mixup'],
                    '201_all_rel': [0, 'two-step Densenet \wo mixup all rel'],
                    '202_all_rel': [1, 'two-step Densenet \wo mixup all rel'],
                    '203_all_rel': [2, 'two-step Densenet \wo mixup all rel'],
                    '204_all_rel': [3, 'two-step Densenet \wo mixup all rel'],
                    '205_all_rel': [4, 'two-step Densenet \wo mixup all rel'],
                    }
    top1_5_avg_df = None
    pr_avg_df = None
    conf_avg_df = None
    top1_5_avg_df_subopt = None
    pr_avg_df_subopt = None
    conf_df_subopt = None

    top1_5_avg_df = pd.DataFrame(columns=['name', 'fold'] + top_column_names)
    pr_avg_df = pd.DataFrame(columns=['name', 'fold'] + pr_column_names)
    conf_dfs = {}
    top1_5_avg_df_subopt = pd.DataFrame(columns=['name', 'fold'] + top_column_names_subopt)
    pr_avg_df_subopt = pd.DataFrame(columns=['name', 'fold'] + pr_column_names_subopt)
    conf_dfs_subopt = {}

    for run_id in results_info:
        if type(run_id) is str and '_' in run_id:
            splitted_run_id = run_id.split('_')
            network_id = splitted_run_id[0]
            postfix = "_" + "_".join(splitted_run_id[1:])
        else:
            network_id = run_id
            postfix = ""
        ind = results_info[run_id][0]
        run_series = pd.Series({'name': results_info[run_id][1], 'fold': results_info[run_id][0]})
        #opt data update
        excel_path = os.path.join(networks_path, str(network_id), 'results', f"predictions_{network_id}{postfix}.xlsx")
        dfs = pd.read_excel(excel_path,sheet_name=sheets_names)
        top_avg = dfs["top1-5"][top_column_names].mean()
        pr_avg = dfs["precision_recall"][pr_column_names].mean()
        top1_5_avg_df.loc[results_info[run_id][1] + '_' + str(results_info[run_id][0])] = run_series.combine_first(top_avg)
        pr_avg_df.loc[results_info[run_id][1] +  '_' + str(results_info[run_id][0])] = run_series.combine_first(pr_avg)
        if results_info[run_id][1] not in conf_dfs:
            conf_dfs[results_info[run_id][1]] = {}
        conf_dfs[results_info[run_id][1]][results_info[run_id][0]] = dfs["conf_matrix"]
        #subopt data update
        excel_path = os.path.join(networks_path, str(network_id), 'results', f"predictions_{network_id}_subopt{postfix}.xlsx")
        dfs = pd.read_excel(excel_path, sheet_name=sheets_names)
        top_avg_subopt = dfs["top1-5"][top_column_names_subopt].mean()
        pr_avg_subopt = dfs["precision_recall"][pr_column_names_subopt].mean()
        top1_5_avg_df_subopt.loc[results_info[run_id][1] +  '_' + str(results_info[run_id][0])] = run_series.combine_first(top_avg_subopt)
        pr_avg_df_subopt.loc[results_info[run_id][1] +  '_' + str(results_info[run_id][0])] = run_series.combine_first(pr_avg_subopt)
    #save all results
    with (pd.ExcelWriter(results_path, engine="openpyxl") as writer):
        top1_5_total_df, pr_total_df, conf_total_dfs, top1_5_total_df_subopt, pr_total_df_subopt =\
            calc_total(top1_5_avg_df, pr_avg_df, conf_dfs, top1_5_avg_df_subopt, pr_avg_df_subopt)

        #write total calculations
        top1_5_total_df.to_excel(writer, sheet_name="top1_5", index=False)
        pr_total_df.to_excel(writer, sheet_name="pr", index=False)
        top1_5_total_df_subopt.to_excel(writer, sheet_name="top1_5_subopt", index=False)
        pr_total_df_subopt.to_excel(writer, sheet_name="pr_subopt", index=False)
        write_dfs_to_sheet(writer, conf_total_dfs, "conf")

        #write original data
        top1_5_avg_df.to_excel(writer, sheet_name="top1_5_folds", index=False)
        pr_avg_df.to_excel(writer, sheet_name="pr_folds", index=False)
        top1_5_avg_df_subopt.to_excel(writer, sheet_name="top_folds_subopt", index=False)
        pr_avg_df_subopt.to_excel(writer, sheet_name="pr_folds_subopt", index=False)








